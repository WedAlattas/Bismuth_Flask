from user import User
from Task import Task
from tempfile import NamedTemporaryFile
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Alignment,Font, Border, Side
import openpyxl
from datetime import datetime
import calendar 
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import io
import boto3

s3 = boto3.resource(
    service_name='s3',
    region_name='us-east-2',
    aws_access_key_id='AKIAY4CEWAN7MPD4ABFC',
    aws_secret_access_key='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'
)
bucket = s3.Bucket('bismuth1')

class Leader(User):

    def __init__ (self, ID):
        self._leaderID = ID

 
    def createTask(self, data, cur): 
        count = cur.execute(''' select * from Project_details where projectName = '{projectName}';'''.format(projectName = data['projectName']))
        print(count)

        if count == 0 : 
            cur.execute('''INSERT INTO Project_details (projectName , programName, departmentName) VALUES ('{projectName}', '{programName}', '{departmentName}');'''.format(projectName = data['projectName'], programName = data['programName'],departmentName = data['departmentName']))
            cur.connection.commit()

        cur.execute('''INSERT INTO Task (taskName , objectives, summary,leaderID, internalSupport,externalSupport, deliverables, projectName) 
        VALUES ('{taskName}', '{objectives}', '{summary}',{leaderID}, '{internalSupport}','{externalSupport}', 
        '{deliverables}','{projectName}');'''.format(taskName = data['taskName'],projectName= data['projectName'], objectives = data['objectives'],summary=data['summary']
        ,internalSupport=data['internalSupport'],externalSupport = data['externalSupport'],deliverables=data['deliverables']
        , leaderID = self._leaderID))
        cur.connection.commit()

        cur.execute('''SELECT MAX(taskID) FROM Task;''')
        taskID = cur.fetchall()
        for geologist in data['AssignedFieldGeologist']:
            cur.execute('''INSERT INTO FieldGeologist_Task (fieldGeologistID,taskID) VALUES
             ({fieldGeologistID},{taskID});'''.format(fieldGeologistID = 
             geologist['userID'] , taskID = taskID[0]['MAX(taskID)']))
            cur.execute('''UPDATE FieldGeologist SET availability= 'False' WHERE fieldGeologistID = {fieldGeologistID}'''.format(fieldGeologistID= int(geologist['userID'])))
            cur.connection.commit()
        return taskID[0]['MAX(taskID)']
       

    def displayTasks(self, status, cur): 
        cur.execute('''SELECT taskName, taskID FROM `Task` INNER JOIN `Project_details` 
        ON `Task`.`projectName` = `Project_details`.`projectName` WHERE 
        `Task`.`status` = '{status}' AND `Task`.`leaderID` ={leaderID};'''.format(status = status, leaderID= self._leaderID))
        tasks = cur.fetchall()
        return tasks

    def editTask(self, cur, data,taskID):

        ## change it from Task.. 
        cur.execute(''' select * from Task where taskID = {taskID};'''.format(taskID = taskID))
        oldTask = cur.fetchall()


        count =cur.execute('''SELECT * FROM Task INNER JOIN Project_details ON Task.projectName = Project_details.projectName WHERE 
           Project_details.projectName ='{projectName}' ;'''.format(projectName= oldTask[0]['projectName']))

        if count > 1 : 
            newProject = cur.execute(''' select * from Project_details where projectName = '{projectName}';'''.format(projectName = data['projectName']))
            if newProject == 0 : 
                cur.execute('''INSERT INTO Project_details (projectName , programName, departmentName) VALUES ('{projectName}', '{programName}', '{departmentName}');'''.format(projectName = data['projectName'], programName = data['programName'],departmentName = data['departmentName']))
                cur.connection.commit()
        elif count == 1:
            cur.execute('''UPDATE Project_details SET projectName = '{projectName}', programName = '{programName}', departmentName = '{departmentName}' where projectName ='{oldPro}';'''.format(projectName = data['projectName'], programName = data['programName'],departmentName = data['departmentName'], oldPro = oldTask[0]['projectName']))
            cur.connection.commit()

        cur.execute('''UPDATE Task SET taskName= '{taskName}', objectives = '{objectives}',summary = '{summary}', internalSupport = '{internalSupport}', externalSupport = '{externalSupport}',deliverables= '{deliverables}', projectName = '{projectName}' WHERE taskID= {taskID};'''.format(taskName = data['taskName'], objectives = data['objectives'],summary=data['summary'],internalSupport=data['internalSupport'],externalSupport = data['externalSupport'],deliverables=data['deliverables'],taskID= taskID,projectName= data['projectName']))
        cur.connection.commit()

    def removeTask(self, cur, taskID): 

        cur.execute(''' select * from FieldGeologist_Task where taskID = {taskID};'''.format(taskID = taskID))
        data = cur.fetchall()

        
        for geologist in data:
            cur.execute('''UPDATE FieldGeologist SET availability= 'True' WHERE fieldGeologistID = {fieldGeologistID}'''.format(fieldGeologistID= int(geologist['fieldGeologistID'])))
            cur.connection.commit()

        cur.execute(''' select * from Task where taskID = {taskID};'''.format(taskID = taskID))
        data = cur.fetchall()
            
        count =cur.execute('''SELECT * FROM Task INNER JOIN Project_details ON Task.projectName = Project_details.projectName WHERE 
           Project_details.projectName ='{projectName}' ;'''.format(projectName= data[0]['projectName']))
        print(count)
        if count == 1 : 
            cur.execute('''DELETE FROM Project_details WHERE projectName ='{projectName}';'''.format(projectName = data[0]['projectName']))
            cur.connection.commit()

        cur.execute('''DELETE FROM FieldGeologist_Task WHERE taskID = {taskID};'''.format(taskID = taskID))
        cur.execute('''DELETE FROM Task WHERE taskID ={taskID};'''.format(taskID = taskID))
        cur.connection.commit()

        


    def generateReport(self, taskID, cur):

        wb = Workbook()
        ws= wb.create_sheet("Input_sheet", 0)
        ws.title = "Input_sheet"

        ws2 = wb.create_sheet("Pictures", 1)
        ws2.title = "Pictures"

        cur.execute('''SELECT * FROM Record where taskID = {taskID};'''.format(taskID = taskID))
        allRecords = cur.fetchall() 
        index = 0
        for record in allRecords:
            yellowFill = PatternFill(start_color='f7d774',
                        end_color='f7d774',
                        fill_type='solid')

            center = Alignment(horizontal='center', vertical='center')
            bold = Font(bold=True)

            thin = Border(left=Side(color='979797', border_style='thin'), 
                                right=Side(color='979797', border_style='thin'), 
                                top=Side(color='979797', border_style='thin'), 
                                bottom=Side(color='979797', border_style='thin'))


            ################## Start First Line ##########################
            cell = 'A' + str(1+index) +':J'+str(1+index)
            cell1 = 'A' + str(1+index)
            ws.merge_cells(cell)
            ws[cell1] = 'Well inventory data recording sheet'
            ws[cell1].alignment = center
            ws[cell1].font= bold


            for cell in ws[cell]:
                cell[0].fill = yellowFill
            
            ################## Start General Site Data  ##########################
            cell = 'A' + str(2+index) +':E'+str(2+index)
            cell1 = 'A' + str(2+index)
            ws.merge_cells(cell)
            ws[cell1] = 'General Site Data'
            ws[cell1].alignment = center


            DrakGray = PatternFill(start_color='9c9c9c',
                            end_color='9c9c9c',
                            fill_type='solid')
            for cell in ws[cell]:
                cell[0].fill = DrakGray

            lightGray = PatternFill(start_color='D5D5D5',
                            end_color='D5D5D5',
                            fill_type='solid')

        
            ws.merge_cells('A' + str(3+index) +':B'+str(3+index))
            ws['A' + str(3+index)] = 'Record No.'
            ws['A' + str(3+index)].fill = lightGray
            ws['A' + str(3+index)].alignment = center




            ws.merge_cells('C' + str(3+index) +':E'+str(3+index))
            ws[ 'C' + str(3+index)] = 'Purpose of the inventory'
            ws[ 'C' + str(3+index)].fill = lightGray
            ws[ 'C' + str(3+index)].alignment = center





            ws.merge_cells('A' + str(4+index) +':B'+str(4+index))
            ws['A' + str(4+index)] = record['recordID']
            ws['A' + str(4+index)].alignment = center


            ws.merge_cells('C' + str(4+index) +':E'+str(4+index))
            ws['C' + str(4+index)] = 'New'
            ws['C' + str(4+index)].alignment = center


            cur.execute('''SELECT name FROM User where userID = {userID};'''.format(userID = record['fieldGeologistID']))
            fieldGeologist = cur.fetchall() 

            ws.merge_cells('A' + str(5+index) +':B'+str(5+index))
            ws['A' + str(5+index)].fill = lightGray
            ws['A' + str(5+index)] = 'Geologist'
            ######################## change #########################
            ws.merge_cells('C' + str(5+index) +':E'+str(5+index))
            ws['C' + str(5+index)] = fieldGeologist[0]['name']
            ws['C' + str(5+index)].alignment = center


            cur.execute('''SELECT * FROM Task where taskID = {taskID};'''.format(taskID = record['taskID']))
            task = cur.fetchall() 


            

            ws.merge_cells('A' + str(6+index) +':B'+str(6+index))
            ws['A' + str(6+index)].fill = lightGray
            ws['A' + str(6+index)] = 'Project Name'
            ######################## change #########################
            ws.merge_cells('C' + str(6+index) +':E'+str(6+index))
            ws['C' + str(6+index)] = task[0]['projectName']
            ws['C' + str(6+index)].alignment = center





            ws.merge_cells('A' + str(7+index) +':B'+str(7+index))
            ws['A' + str(7+index)].fill = lightGray
            ws['A' + str(7+index)] = 'Task Name'
            ## change 
            ws.merge_cells('C' + str(7+index) +':E'+str(7+index))
            ws['C' + str(7+index)] = task[0]['taskName']
            ws['C' + str(7+index)].alignment = center


            date_time_obj = datetime.strptime(record['date'], '%Y-%m-%d')


            ws.merge_cells('A' + str(8+index) +':B'+str(8+index))
            ws['A' + str(8+index)].fill = lightGray
            ws['A' + str(8+index)] = 'Date of Inventory'
            ## change 
            ws['C' + str(8+index)] = date_time_obj.day
            ws['C' + str(8+index)].alignment = center

            ws['D' + str(8+index)] = date_time_obj.month
            ws['D' + str(8+index)].alignment = center

            ws['E' + str(8+index)] = date_time_obj.year
            ws['E' + str(8+index)].alignment = center






            ws.merge_cells('A' + str(9+index) +':B'+str(9+index))
            ws['A' + str(9+index)].fill = lightGray
            ws['A' + str(9+index)] = 'Day'
            ## change 
            ws.merge_cells('C' + str(9+index) +':E'+str(9+index))
            day = calendar.weekday(date_time_obj.year, date_time_obj.month, date_time_obj.day)
            ws['C' + str(9+index)] =  calendar.day_name[day]
            ws['C' + str(9+index)].alignment = center
            ws.merge_cells('A' + str(10+index) +':B'+str(10+index))
            ws['A' + str(10+index)].fill = lightGray
            ws['A' + str(10+index)] = 'City / Town'
            ## change 
            ws.merge_cells('C' + str(10+index) +':E'+str(10+index))
            ws['C' + str(10+index)] = record['city']
            ws['C' + str(10+index)].alignment = center
            ws.merge_cells('A' + str(11+index) +':B'+str(11+index))
            ws['A' + str(11+index)].fill = lightGray
            ws['A' + str(11+index)] = 'Site Name'
            ## change 
            ws.merge_cells('C' + str(11+index) +':E'+str(11+index))
            ws['C' + str(11+index)] = record['siteName']
            ws['C' + str(11+index)].alignment = center

            ws.merge_cells('A' + str(12+index) +':B'+str(12+index))
            ws['A' + str(12+index) ].fill = lightGray
            ws['A' + str(12+index) ] = 'Stream'
            ## change 
            ws.merge_cells('C' + str(12+index) +':E'+str(12+index))
            ws['C' + str(12+index)] = record['stream']
            ws['C' + str(12+index)].alignment = center





            ws.merge_cells('A' + str(13+index) +':A'+str(14+index))
            ws['A' + str(13+index) ].fill = lightGray
            ws['A' + str(13+index) ] = 'Coordinates'
            ws['A' + str(13+index) ].alignment = center
            ## change 
            ws['B' + str(13+index) ] = 'Latitude'
            ws['B' + str(13+index) ].fill = lightGray
            ws['B' + str(14+index) ] = 'Longitude'
            ws['B' + str(14+index) ].fill = lightGray
            ws.merge_cells('C' + str(13+index) +':E'+str(13+index))
            ws['C' + str(13+index) ]= record['latitude']
            ws['C' + str(13+index) ].alignment = center
            ws.merge_cells('C' + str(14+index) +':E'+str(14+index))
            ws['C' + str(14+index) ]= record['longitude']
            ws['C' + str(14+index) ].alignment = center

            #length 28
            ws.merge_cells('F' + str(2+index) +':F'+str(14+index))
            ws.merge_cells('G' + str(2+index) +':J'+str(2+index))
            ws['G' + str(2+index) ] = 'Field Analysis '
            currentCell = ws['G' + str(2+index) ] #or currentCell = ws['A1']
            currentCell.alignment = Alignment(horizontal='center')

            redFill = PatternFill(start_color='9c9c9c',
                            end_color='9c9c9c',
                            fill_type='solid')
            for cell in ws['G' + str(2+index) +':J'+str(2+index)]:
                cell[0].fill = redFill



            ws['G' + str(3+index)]='EC (uS/cm)'
            ws['G' + str(3+index)].alignment= center
            ws['G' + str(4+index)]=record['EC']
            ws['G' + str(4+index)].alignment= center


            ws['H' + str(3+index) ]='pH'
            ws['H' + str(3+index) ].alignment= center
            ws['H' + str(4+index) ]=record['pH']
            ws['H' + str(4+index) ].alignment= center



            ws['I' + str(3+index) ]='Temp'
            ws['I' + str(3+index) ].alignment= center
            ws['I' + str(4+index) ]=record['temp']
            ws['I' + str(4+index) ].alignment= center





            ws['J' + str(3+index) ]='EH'
            ws['J' + str(3+index) ].alignment= center
            ws['J' + str(4+index) ]=record['EH']
            ws['J' + str(4+index) ].alignment= center




            side = openpyxl.styles.Side(border_style='thin',color='FFFFFF')
            xxx = openpyxl.styles.borders.Border(
                left=side, 
                right=side, 
            )



            ws.merge_cells('G' + str(5+index) +':J'+str(5+index))
            ws['G' + str(5+index) ] = 'Remarks'
            ws['G' + str(5+index) ].alignment = center
            ws['G' + str(5+index) ].fill = DrakGray

            ws.merge_cells('G' + str(6+index) +':J'+str(14+index))
            ws['G' + str(6+index) ].alignment = center
            ws['G' + str(6+index) ] = record['remark']


            ws.merge_cells('A' + str(15+index) +':J'+str(15+index))
            index = index+15
        side = openpyxl.styles.Side(border_style='thin',color='979797')
        no_border = openpyxl.styles.borders.Border(
                left=side, 
                right=side, 
                top=side, 
                bottom=side,
        )

            # Loop through all cells in all worksheets
        for sheet in wb.worksheets:
            for row in sheet:
                for cell in row:
                        # Apply colorless and borderless styles
                    cell.border = no_border
        

        center = Alignment(horizontal='center', vertical='center')
        ws2.title = 'Pictures'

        lightGray = PatternFill(start_color='D5D5D5',
                            end_color='D5D5D5',
                            fill_type='solid')

        DrakGray = PatternFill(start_color='9c9c9c',
                            end_color='9c9c9c',
                            fill_type='solid')


        row = 0
        column = 0
        for record in allRecords: 
            ws2['A'+str(1+row)] = 'Record No.'
            ws2['A'+str(1+row)].alignment = center
            ws2['B'+str(1+row)] = 'Image ID'
            ws2['B'+str(1+row)].alignment = center

            ws2['A'+str(1+row)].fill = DrakGray
            ws2['B'+str(1+row)].fill = DrakGray

            ws2.merge_cells('A'+str(2+row)+':A'+str(3+row)) 

            ws2['A'+str(2+row)] = record['recordID']
            ws2['A'+str(2+row)].alignment = center

            ws2['B' +str(2+row)] = 'Label'
            ws2['B'+str(2+row)].alignment = center


            ws2['B'+str(3+row)] = 'Image'
            ws2['B'+str(3+row)].alignment = center

            ws2['B'+str(2+row)].fill = lightGray

            ws2['B'+str(3+row)].fill = lightGray

            cur.execute('''SELECT * FROM Image where recordID = {recordID};'''.format(recordID = record['recordID']))
            images = cur.fetchall() 
            ws2.row_dimensions[3+row].height =  221
            for image in images: 

                ws2.column_dimensions[str(chr(67+column))].width = 35

                ws2[str(chr(67+column))+str(1+row)] = image['imageID']
                ws2[str(chr(67+column))+str(1+row)].alignment = center
                ws2[str(chr(67+column))+str(2+row)] = image['imageType']
                ws2[str(chr(67+column))+str(2+row)].alignment = center


                image = bucket.Object(image['imagePath']).get().get('Body')
                imageStream = io.BytesIO(image.read())
                logo = PILImage.open(imageStream)
                logo = XLImage(logo)
                logo.height = 294
                logo.width = 280


                ws2.add_image(logo, str(chr(67+column))+str(3+row))
                column= column+1
            row = row+4
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=column+2)
            column = 0

        side = openpyxl.styles.Side(border_style='thin',color='979797')
        no_border = openpyxl.styles.borders.Border(
                left=side, 
                right=side, 
                top=side, 
                bottom=side,
            )

    # Loop through all cells in all worksheets
        for sheet in wb.worksheets:
            for row in sheet:
                for cell in row:
                    # Apply colorless and borderless styles
                    cell.border = no_border


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            

        bucket.Object("Reports/{taskID}.xlsx".format(taskID = taskID)).put(Body=stream)
        return "Reports/{taskID}.xlsx".format(taskID = taskID)



    @property
    def leaderID(self):
        return self._leaderID

    # define setter function for leaderID property
    @leaderID.setter 
    def leaderID(self, leaderID):
        self._leaderID= leaderID

    





